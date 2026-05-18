"""ייצוא תוצאות חיפוש ומק״ט+ספקים לקובץ Excel."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from db_service import ITEM_LIST_COLUMNS, VARIANT_LABELS

META_SHEET = "מידע"
MAKT_SHEET = "מקטים"
VARIANTS_SHEET = "וריאנטים"
SUPPLIERS_SHEET = "ספקים"

SUPPLIER_EXPORT_COLUMNS = [
    ("מספר ספק", "מספר ספק"),
    ("שם ספק", "שם ספק"),
    ("יישוב קליניקה", "יישוב קליניקה"),
    ("טלפון", "_phone"),
    ("אזור", "אזור"),
    ("האם בתוקף", "האם בתוקף"),
    ("מחיר הסכם", "מחיר הסכם"),
]

PHONE_KEYS = ("נייד ספק", "טלפון עבודה ספק", "נייח ספק")


def _now_he() -> str:
    return datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M")


def _format_phone_display(raw: Any) -> str:
    text = str(raw or "").strip()
    if not text or text in ("—", "לא מוגדר"):
        return ""
    if text.startswith("0"):
        return text
    digits = "".join(c for c in text if c.isdigit())
    if not digits:
        return text
    if digits.startswith("972"):
        digits = "0" + digits[3:]
    if not digits.startswith("0") and len(digits) >= 7:
        area = digits[0]
        if (len(digits) == 9 and area == "5") or (
            8 <= len(digits) <= 9 and area in "23489"
        ):
            return f"0{text}" if text and text[0].isdigit() else f"0{digits}"
    return text


def _supplier_phone(s: dict[str, Any]) -> str:
    for key in PHONE_KEYS:
        val = s.get(key)
        if val:
            return _format_phone_display(val)
    return ""


def _amount_range(variants: list[dict[str, Any]]) -> str:
    nums: list[float] = []
    for v in variants:
        raw = v.get("סכום")
        if raw is None:
            continue
        try:
            nums.append(float(str(raw).replace(",", "")))
        except ValueError:
            pass
    if not nums:
        return ""
    lo, hi = min(nums), max(nums)
    if lo == hi:
        return str(lo)
    return f"{lo} – {hi}"


def _write_meta_rows(ws, rows: list[tuple[str, Any]]) -> None:
    ws.sheet_view.rightToLeft = True
    bold = Font(bold=True)
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value if value is not None else "")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    ws.sheet_view.rightToLeft = True
    header_font = Font(bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for r in range(2, min(len(rows) + 2, 52)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 80))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def _workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_search_export(
    *,
    query: str,
    match: str,
    field: str,
    groups: list[dict[str, Any]],
    items: list[dict[str, Any]],
) -> bytes:
    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = META_SHEET
    _write_meta_rows(
        ws_meta,
        [
            ("סוג ייצוא", "תוצאות חיפוש"),
            ("תאריך ייצוא", _now_he()),
            ("שאילתה", query),
            ("התאמה", match),
            ("שדה", field),
            ("מספר מקטים", len(groups)),
            ("מספר וריאנטים", len(items)),
        ],
    )

    ws_makt = wb.create_sheet(MAKT_SHEET)
    makt_headers = [
        'מק"ט',
        "תיאור פריט",
        "מספר וריאנטים",
        "מספר ספקים",
        "טווח סכום",
        "סוג זכאי (ראשון)",
    ]
    makt_rows: list[list[Any]] = []
    for g in groups:
        variants = g.get("variants") or []
        makt_rows.append(
            [
                g.get('מק"ט'),
                g.get("תיאור פריט") or (variants[0].get("תיאור פריט") if variants else ""),
                g.get("variant_count") or len(variants),
                g.get("supplier_count", 0),
                _amount_range(variants),
                variants[0].get("סוג זכאי") if variants else "",
            ]
        )
    _write_table(ws_makt, makt_headers, makt_rows)

    ws_var = wb.create_sheet(VARIANTS_SHEET)
    var_headers = list(ITEM_LIST_COLUMNS)
    var_rows = [[item.get(c) for c in ITEM_LIST_COLUMNS] for item in items]
    _write_table(ws_var, var_headers, var_rows)

    return _workbook_to_bytes(wb)


def build_makt_export(
    *,
    makt: str,
    variants: list[dict[str, Any]],
    suppliers: list[dict[str, Any]],
    selected_entity_id: str | None = None,
    selected_variant: dict[str, Any] | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = META_SHEET

    desc = ""
    if variants:
        desc = str(variants[0].get("תיאור פריט") or "")

    meta_rows: list[tuple[str, Any]] = [
        ("סוג ייצוא", "מק״ט + ספקים מורשים"),
        ("תאריך ייצוא", _now_he()),
        ('מק"ט', makt),
        ("תיאור פריט", desc),
        ("הערה", "ספקים מורשים ברמת מק״ט — זהים לכל הוריאנטים"),
        ("מספר וריאנטים", len(variants)),
        ("מספר ספקים", len(suppliers)),
    ]
    if selected_entity_id:
        meta_rows.append(("וריאנט נבחר (entity_id)", selected_entity_id))
        if selected_variant:
            for key, _ in VARIANT_LABELS:
                val = selected_variant.get(key)
                if val is not None and str(val) not in ("", "לא מוגדר", "0", "0.0"):
                    meta_rows.append((key, val))

    _write_meta_rows(ws_meta, meta_rows)

    ws_var = wb.create_sheet(VARIANTS_SHEET)
    var_headers = list(ITEM_LIST_COLUMNS)
    var_rows = [[v.get(c) for c in ITEM_LIST_COLUMNS] for v in variants]
    _write_table(ws_var, var_headers, var_rows)

    ws_sup = wb.create_sheet(SUPPLIERS_SHEET)
    sup_headers = [h for h, _ in SUPPLIER_EXPORT_COLUMNS]
    sup_rows: list[list[Any]] = []
    for s in suppliers:
        row: list[Any] = []
        for _title, key in SUPPLIER_EXPORT_COLUMNS:
            if key == "_phone":
                row.append(_supplier_phone(s))
            else:
                row.append(s.get(key))
        sup_rows.append(row)
    _write_table(ws_sup, sup_headers, sup_rows)

    return _workbook_to_bytes(wb)


def build_ai_search_export(payload: dict[str, Any]) -> bytes:
    """ייצוא תוצאות חיפוש חכם — מקטים, וריאנטים וספקים (כולל קרבה)."""
    query = str(payload.get("query") or "")
    parsed = payload.get("parsed") or {}
    user_loc = payload.get("user_location") or ""
    results: list[dict[str, Any]] = payload.get("results") or []

    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = META_SHEET
    _write_meta_rows(
        ws_meta,
        [
            ("סוג ייצוא", "חיפוש חכם"),
            ("תאריך ייצוא", _now_he()),
            ("שאילתה", query),
            ("מנוע", payload.get("engine") or "local"),
            ("מיקום משתמש", user_loc or "—"),
            ("הסבר", parsed.get("explanation") or ""),
            ("ביטוי חיפוש", parsed.get("search_phrase") or ""),
            ("מספר מקטים", len(results)),
        ],
    )

    ws_makt = wb.create_sheet(MAKT_SHEET)
    makt_headers = [
        'מק"ט',
        "תיאור פריט",
        "מספר וריאנטים",
        "מספר ספקים",
        "ספק הכי קרוב",
        "יישוב ספק קרוב",
        "טווח סכום",
    ]
    makt_rows: list[list[Any]] = []
    all_variants: list[list[Any]] = []
    all_suppliers: list[list[Any]] = []

    for r in results:
        makt = r.get('מק"ט')
        variants = r.get("variants") or []
        suppliers = r.get("suppliers") or []
        nearest = r.get("nearest_supplier") or {}
        if not nearest:
            nearest = next((s for s in suppliers if s.get("is_nearest")), {})

        makt_rows.append(
            [
                makt,
                r.get("תיאור פריט") or "",
                r.get("variant_count") or len(variants),
                r.get("supplier_count") or len(suppliers),
                nearest.get("שם ספק") or "",
                nearest.get("יישוב קליניקה") or "",
                _amount_range(variants),
            ]
        )

        for v in variants:
            all_variants.append([makt] + [v.get(c) for c in ITEM_LIST_COLUMNS])

        for s in suppliers:
            proximity = "הכי קרוב" if s.get("is_nearest") else (s.get("proximity_label") or "")
            all_suppliers.append(
                [
                    makt,
                    proximity,
                    s.get("שם ספק"),
                    s.get("יישוב קליניקה"),
                    _supplier_phone(s),
                    s.get("אזור"),
                    s.get("האם בתוקף"),
                    s.get("מחיר הסכם"),
                ]
            )

    _write_table(ws_makt, makt_headers, makt_rows)

    ws_var = wb.create_sheet(VARIANTS_SHEET)
    _write_table(ws_var, ['מק"ט'] + list(ITEM_LIST_COLUMNS), all_variants)

    ws_sup = wb.create_sheet(SUPPLIERS_SHEET)
    sup_headers = [
        'מק"ט',
        "קרבה",
        "שם ספק",
        "יישוב",
        "טלפון",
        "אזור",
        "בתוקף",
        "מחיר הסכם",
    ]
    _write_table(ws_sup, sup_headers, all_suppliers)

    return _workbook_to_bytes(wb)
